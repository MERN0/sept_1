"""Create a test Excel file with proper sheet structure for Node 2 testing"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

output_dir = "files/test@tataelxsi.co.in/swe6_5/V1.0/SWE6/input"
os.makedirs(output_dir, exist_ok=True)

output_file = os.path.join(output_dir, "reqs_to_use.xlsx")

# Create workbook
wb = openpyxl.Workbook()

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

print("[LOG] Creating Index sheet...")
# Create Index sheet
index_ws = wb.create_sheet("Index")
index_ws.append(["Feature Number", "Feature Name", "Feature Group"])
index_ws.append(["019", "Drive Motor Control", "Power Management"])
index_ws.append(["020", "Braking System", "Safety"])
index_ws.append(["021", "Steering Control", "Safety"])

# Create Master Comm Matrix (CAN) sheet
print("[LOG] Creating Master Comm Matrix (CAN) sheet...")
comm_ws = wb.create_sheet("Master Comm Matrix (CAN)")

# Headers: Start with some columns, then feature columns
headers = [
    "Message ID", "Signal Name", "Signal Type", "Sender", "Receiver",
    "Start Bit", "Length", "Scale", "Offset", "Min", "Max",
    "Unit", "Comment", "A", "B", "C", "D", "E",  # Some filler columns
    "019", "020", "021", "022"  # Feature columns
]

for col_idx, header in enumerate(headers, 1):
    cell = comm_ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

# Add sample signal data
signals = [
    [100, "Drv1_Rx1_CmdSpd", "Message", "ECU1", "Motor1", 0, 16, 0.1, 0, 0, 100, "RPM", "Speed command"],
    [101, "Drv1_Rx1_CmdTorque", "Message", "ECU1", "Motor1", 16, 16, 0.5, -500, -500, 500, "Nm", "Torque command"],
    [102, "Drv1_Rx1_CmdDec", "Message", "ECU1", "Motor1", 32, 8, 1, 0, 0, 100, "%", "Deceleration"],
    [103, "Status_PowerOn", "Signal", "ECU1", "Monitor", 40, 1, 1, 0, 0, 1, "Boolean", "Power status"],
    [104, "Status_Error", "Signal", "ECU1", "Monitor", 41, 1, 1, 0, 0, 1, "Boolean", "Error flag"],
]

for row_idx, signal_data in enumerate(signals, 2):
    for col_idx, value in enumerate(signal_data, 1):
        comm_ws.cell(row=row_idx, column=col_idx, value=value)

# Add markers to feature columns
# Feature 019 column is at index that corresponds to "019" header
feature_019_col = headers.index("019") + 1
feature_020_col = headers.index("020") + 1
feature_021_col = headers.index("021") + 1

# Mark some signals as using these features
comm_ws.cell(row=2, column=feature_019_col, value="x")      # Drv1_Rx1_CmdSpd uses feature 019
comm_ws.cell(row=3, column=feature_019_col, value="〇")      # Drv1_Rx1_CmdTorque uses feature 019
comm_ws.cell(row=4, column=feature_019_col, value="x")      # Drv1_Rx1_CmdDec uses feature 019
comm_ws.cell(row=2, column=feature_020_col, value="x")      # Drv1_Rx1_CmdSpd uses feature 020
comm_ws.cell(row=5, column=feature_021_col, value="〇")      # Status_Error uses feature 021

print("[LOG] Creating requirements sheet (005)...")
# Create requirements sheet (005)
req_ws = wb.create_sheet("005")
req_ws.append([
    "ID", "Description", "Verification Criteria", "Status", "Comments"
])
req_ws.append([
    "REQ_019_001",
    "Motor speed command must be transmitted",
    "The speed command signal should be sent every 100ms with accurate RPM values",
    "Functional",
    "Required for motor control"
])
req_ws.append([
    "REQ_019_002",
    "Motor torque must be controllable",
    "Torque command should be applied within 50ms of reception and respect min/max limits",
    "Functional",
    "Safety critical"
])
req_ws.append([
    "REQ_020_001",
    "Braking system must be responsive",
    "Brake command must be executed within 200ms, with proportional force application",
    "Functional",
    "Safety critical"
])

# Adjust column widths
for ws in wb.sheetnames:
    sheet = wb[ws]
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = min(adjusted_width, 50)

# Save workbook
wb.save(output_file)
print(f"[SUCCESS] Test Excel file created: {output_file}")
print(f"\nSheets created:")
for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f"  - {sheet}: {len(list(ws.rows))} rows")

print(f"\nMaster Comm Matrix (CAN) sheet structure:")
comm_ws = wb["Master Comm Matrix (CAN)"]
print(f"  Columns: {list(comm_ws[1])[:5]}... (showing first 5)")
print(f"  Feature 019 column: {get_column_letter(feature_019_col)}")
print(f"  Feature 020 column: {get_column_letter(feature_020_col)}")
print(f"  Feature 021 column: {get_column_letter(feature_021_col)}")
