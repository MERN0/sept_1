"""
Excel writer for the SYS5 output workbook - structure matches the reference
implementation in the work_28 repo's xlsx_writer.py (same tab layout, same
Test Cases merge strategy), adapted to this pipeline's own state shapes
(plain dicts from Nodes 1-9, not work_28's pydantic schema).

Sheets written: Cover Page, Test Pattern, Item List, Configurable Parameters,
Test Cases.

Once Node 7 (Generate) produces real content, each entry in
state["test_cases"][req_id]["generated_output"] should be shaped as:
{
    "test_case_id": str,
    "feature": str,
    "variant": Optional[str],
    "requirement_ids": list[str],
    "priority": Optional[str],
    "mode_of_execution": str,          # e.g. "Automated"
    "description": str,
    "status": "clean" | "flagged",
    "flag_reason": Optional[str],
    "remarks_summary": Optional[str],  # compact summary for Item List
    "steps": [
        {
            "step_no": int,
            "phase": "PRECONDITION" | "ACTION" | "POSTCONDITION",
            "step_text": str,
            "parameter_settings": Optional[str],
            "units": Optional[str],
            "expected_value": Optional[str],
            "units2": Optional[str],
            "whether_execute": str,    # "Yes"/"No"
            "remarks": Optional[str],
        },
        ...
    ],
}
This writer never crashes on missing/placeholder fields - everything is
read with .get(...) and blanks are written for anything absent, since
Nodes 7-9 only produce this shape once the Generate/Validate/Correct
prompts (prompts.py) are written.
"""

import os
import re
from datetime import datetime, timezone
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_HEADER_FONT = Font(bold=True)
_TOP_ALIGN = Alignment(vertical="top", wrap_text=True)

_TEST_CASE_COLUMNS = [
    "TestCase ID", "Feature", "Variant", "Requirement IDs", "Priority", "Mode of Execution",
    "Test case Description", "Test Phase", "Step", "Test steps", "Parameter Settings", "Units",
    "Expected Value", "Units", "Whether to execute the command", "Remarks",
]

_ITEM_LIST_COLUMNS = [
    "Testcase No", "Test type", "Feature Name", "Test Case ID", "Variant", "Execution Required", "Remarks"
]


def _style_header_row(ws: Worksheet, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL


def _write_cover_page(wb: Workbook, state: Dict[str, Any], config: Dict[str, Any]) -> None:
    ws = wb.create_sheet("Cover Page")
    ws.append(["System Qualification Test Specification"])
    ws.append([])
    ws.append(["Project", config.get("project_name", "")])
    ws.append(["Version", config.get("version", "")])
    ws.append(["Feature", f"{config.get('feature_id', '')} - {config.get('feature_name', '')}".strip(" -")])
    ws.append(["Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")])
    ws.append(["Requirements processed", len(state.get("requirements", []))])
    test_cases = state.get("test_cases", {})
    ws.append(["Test cases tracked", len(test_cases)])
    flagged = sum(1 for tc in test_cases.values() if (tc.get("generated_output") or {}).get("status") == "flagged")
    ws.append(["Test cases flagged for review", flagged])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60


def _write_test_pattern(wb: Workbook, state: Dict[str, Any]) -> None:
    """
    This pipeline's test_patterns[req_id] shape is {"test_cases": [...],
    "factors": {...}, "summary": ...} (Node 1) rather than work_28's
    TestPatternRow list, so columns are derived from the union of
    preconditions/actions field names actually present instead of a fixed
    fixed/variable factor split.
    """
    ws = wb.create_sheet("Test Pattern")
    patterns = state.get("test_patterns", {})

    precondition_names: List[str] = []
    action_names: List[str] = []
    for pattern in patterns.values():
        for tc in pattern.get("test_cases", []):
            for name in tc.get("preconditions", {}):
                if name not in precondition_names:
                    precondition_names.append(name)
            for name in tc.get("actions", {}):
                if name not in action_names:
                    action_names.append(name)

    header = ["Requirement ID", "Test Case No."] + precondition_names + action_names + ["Expected Result"]
    ws.append(header)
    _style_header_row(ws, 1, len(header))

    for req_id, pattern in patterns.items():
        for tc in pattern.get("test_cases", []):
            row = [req_id, tc.get("test_case_no", "")]
            row += [tc.get("preconditions", {}).get(name, "") for name in precondition_names]
            row += [tc.get("actions", {}).get(name, "") for name in action_names]
            row.append(tc.get("expected_result", ""))
            ws.append(row)

    for col in range(1, len(header) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18


def _write_item_list(wb: Workbook, test_cases: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("Item List")
    ws.append(_ITEM_LIST_COLUMNS)
    _style_header_row(ws, 1, len(_ITEM_LIST_COLUMNS))

    for i, tc in enumerate(test_cases, start=1):
        ws.append([
            i,
            tc.get("test_type", "Normal_system"),
            tc.get("feature", ""),
            tc.get("test_case_id", ""),
            tc.get("variant", ""),
            "Yes" if tc.get("steps") else "",
            tc.get("remarks_summary", ""),
        ])

    widths = [12, 16, 20, 16, 10, 16, 40]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width


def _write_configurable_parameters(wb: Workbook, feature_details: Dict[str, Any]) -> None:
    """
    Best-effort adaptation: work_28 sources this from a dedicated
    app_param_valid list with per-variant columns (A1/A2/B/C1/...). This
    pipeline's Node 4 instead stores App Parameter rows inside
    feature_details keyed by f"{sheet_name}_{header_value}" with dynamic
    column names, so per-variant columns aren't separately resolved here -
    only Parameter Name + a best-guess Description column are populated.
    """
    ws = wb.create_sheet("Configurable Parameters")
    header = ["Parameter Name", "Description", "A1", "A2", "B", "C1", "C2", "C3", "D1", "D2"]
    ws.append(header)
    _style_header_row(ws, 1, len(header))

    seen = set()
    for entry in feature_details.values():
        if not isinstance(entry, dict) or "header_value" not in entry:
            continue  # not a Node 4 App Parameter entry
        name = str(entry.get("header_value", "")).strip()
        if not name or name in seen:
            continue
        seen.add(name)
        description = next(
            (str(v) for k, v in entry.items() if "desc" in str(k).lower() and v), ""
        )
        ws.append([name, description, "", "", "", "", "", "", "", ""])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60


def _write_test_cases(wb: Workbook, test_cases: List[Dict[str, Any]]) -> None:
    """
    One row per step. Columns 1-7 (metadata) and the Test Phase column (8)
    are merged vertically across a test case's / phase's rows rather than
    repeated on every row. One blank row is left between consecutive test
    cases. A test case with no steps yet (nothing generated) is skipped.
    """
    ws = wb.create_sheet("Test Cases")
    ws.append(_TEST_CASE_COLUMNS)
    _style_header_row(ws, 1, len(_TEST_CASE_COLUMNS))

    row_idx = 2
    for tc in test_cases:
        steps = tc.get("steps") or []
        if not steps:
            continue
        block_start = row_idx

        phase_start = row_idx
        current_phase = steps[0].get("phase", "")
        for step in steps:
            if step.get("phase", "") != current_phase:
                ws.merge_cells(start_row=phase_start, start_column=8, end_row=row_idx - 1, end_column=8)
                ws.cell(row=phase_start, column=8, value=current_phase)
                current_phase = step.get("phase", "")
                phase_start = row_idx

            ws.cell(row=row_idx, column=9, value=step.get("step_no", ""))
            ws.cell(row=row_idx, column=10, value=step.get("step_text", ""))
            ws.cell(row=row_idx, column=11, value=step.get("parameter_settings", "") or "")
            ws.cell(row=row_idx, column=12, value=step.get("units", "") or "")
            ws.cell(row=row_idx, column=13, value=step.get("expected_value", "") or "")
            ws.cell(row=row_idx, column=14, value=step.get("units2", "") or "")
            ws.cell(row=row_idx, column=15, value=step.get("whether_execute", "Yes"))
            ws.cell(row=row_idx, column=16, value=step.get("remarks", "") or "")
            row_idx += 1

        ws.merge_cells(start_row=phase_start, start_column=8, end_row=row_idx - 1, end_column=8)
        ws.cell(row=phase_start, column=8, value=current_phase)

        block_end = row_idx - 1
        if block_end > block_start:
            for col in range(1, 8):
                ws.merge_cells(start_row=block_start, start_column=col, end_row=block_end, end_column=col)
        ws.cell(row=block_start, column=1, value=tc.get("test_case_id", ""))
        ws.cell(row=block_start, column=2, value=tc.get("feature", ""))
        ws.cell(row=block_start, column=3, value=tc.get("variant", "") or "")
        ws.cell(row=block_start, column=4, value=", ".join(tc.get("requirement_ids", [])))
        ws.cell(row=block_start, column=5, value=tc.get("priority", "") or "")
        ws.cell(row=block_start, column=6, value=tc.get("mode_of_execution", "Automated"))
        description = tc.get("description", "")
        if tc.get("status") == "flagged" and tc.get("flag_reason"):
            description = f"{description}\n[FLAGGED FOR REVIEW: {tc['flag_reason']}]"
        ws.cell(row=block_start, column=7, value=description).alignment = _TOP_ALIGN

        row_idx += 1  # one blank row between consecutive test cases

    widths = [16, 16, 10, 20, 10, 16, 40, 14, 6, 40, 18, 8, 18, 8, 12, 30]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width


def _feature_name_for(req_id: str, feature_index: Dict[str, Any], fallback: str) -> str:
    """
    The Feature column must be the real Feature Name from the Index sheet
    (state["feature_index"], Node 1), never the LLM-generated value - only
    fall back to whatever Generate produced if the requirement's feature
    number can't be resolved in the index at all (e.g. no Index sheet).

    Every miss is logged with the reason (no 3-digit number in req_id, vs.
    a number that didn't match any Index sheet key) - Node 1 already logs
    what it parsed out of the Index sheet, so pairing that with what this
    lookup saw for each req_id makes a mismatch (e.g. a req_id whose digits
    don't line up with the Index sheet's Feature Number) immediately
    diagnosable from the run output instead of a silently blank/fallback
    Feature column.
    """
    match = re.search(r'(\d{3})', req_id or "")
    if not match:
        print(f"[LOG] Feature lookup for '{req_id}': no 3-digit feature number found in the requirement id - "
              f"falling back to Generate's own feature value\n")
        return fallback

    entry = feature_index.get(match.group(1))
    if entry and entry.get("feature_name"):
        return entry["feature_name"]

    print(f"[LOG] Feature lookup for '{req_id}': extracted feature number '{match.group(1)}' but it's not a key "
          f"in the Index sheet's feature_index ({sorted(feature_index.keys())}) - falling back to Generate's own "
          f"feature value\n")
    return fallback


def _normalize_test_cases(state: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Build the flat list write_test_cases_workbook's sheet writers expect,
    from state["test_cases"] (dict keyed by "{req_id}_{pattern_no}", one
    entry per generated test case - a requirement can have several, one per
    test pattern entry) + state["requirements"]. Falls back to the
    requirement's own description/req_id when nothing has been generated
    yet, so Item List / Test Cases still get a usable row.

    The dict key is no longer the req_id itself (Node 7 makes one test case
    per test pattern entry, not one per requirement), so the real req_id is
    read from entry["req_id"] rather than parsed back out of the key.
    """
    requirements_by_id = {req.get("req_id"): req for req in state.get("requirements", [])}
    feature_index = state.get("feature_index", {})
    normalized = []
    for key, entry in state.get("test_cases", {}).items():
        generated = entry.get("generated_output") or {}
        req_id = entry.get("req_id", key)
        requirement = requirements_by_id.get(req_id, {})
        normalized.append({
            "test_case_id": generated.get("test_case_id") or entry.get("test_case_id", key),
            "feature": _feature_name_for(req_id, feature_index, generated.get("feature", "")),
            "variant": generated.get("variant"),
            "requirement_ids": generated.get("requirement_ids", [req_id]),
            "priority": generated.get("priority"),
            "mode_of_execution": generated.get("mode_of_execution", "Automated"),
            "description": generated.get("description", requirement.get("data", {}).get("Description", "")),
            "status": generated.get("status", "clean"),
            "flag_reason": generated.get("flag_reason"),
            "remarks_summary": generated.get("remarks_summary"),
            "test_type": generated.get("test_type", "Normal_system"),
            "steps": generated.get("steps", []),
        })
    return normalized


def write_test_cases_workbook(output_path: str, state: Dict[str, Any]) -> None:
    """
    Build the full SYS5 output workbook (Cover Page, Test Pattern, Item
    List, Configurable Parameters, Test Cases) from the pipeline's final
    state dict and save it to output_path.
    """
    config = state.get("config", {})
    test_cases = _normalize_test_cases(state)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)
    _write_cover_page(wb, state, config)
    _write_test_pattern(wb, state)
    _write_item_list(wb, test_cases)
    _write_configurable_parameters(wb, state.get("feature_details", {}))
    _write_test_cases(wb, test_cases)

    wb.save(output_path)
