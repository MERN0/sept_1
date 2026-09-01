"""Node 1: Requirements Extraction from Excel"""

import json
import os
import sys
from typing import Dict, Any
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Setup path for imports
_workspace_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../../../.."))
if _workspace_root not in sys.path:
    sys.path.insert(0, _workspace_root)

try:
    from ..state import SYS5State  # type: ignore
    from ..utils import (resolve_path, ensure_directory_exists, extract_verification_criteria,  # type: ignore
                         prepare_test_pattern_prompt, parse_test_patterns_json, format_test_patterns_for_excel)
    from ..config import is_functional_requirement, KEYWORD_MATCHING_CONFIG, FUNCTIONAL_REQ_KEYWORDS, get_llm  # type: ignore
except ImportError:
    from backend.app.core.artifacts.system.sys5.state import SYS5State
    from backend.app.core.artifacts.system.sys5.utils import (
        resolve_path, ensure_directory_exists, extract_verification_criteria,
        prepare_test_pattern_prompt, parse_test_patterns_json, format_test_patterns_for_excel
    )
    from backend.app.core.artifacts.system.sys5.config import (
        is_functional_requirement, KEYWORD_MATCHING_CONFIG, FUNCTIONAL_REQ_KEYWORDS, get_llm
    )


class Node1ExtractRequirements:
    """
    Node 1: Extract functional requirements from Excel file

    Process:
    1. Resolve input/output paths (relative or absolute)
    2. Read Excel file
    3. Scan for rows containing "Functional requirements"
    4. Extract matching rows with all data
    5. Save results to JSON file
    """

    @staticmethod
    def execute(state: SYS5State) -> SYS5State:
        """
        Execute Node 1 - Requirements Extraction

        Args:
            state: Current workflow state

        Returns:
            Updated state with extracted requirements
        """
        print(f"\n{'='*80}")
        print("NODE 1: REQUIREMENTS EXTRACTION")
        print(f"{'='*80}\n")

        config = state["config"]
        requirements = []
        errors = []

        # Extract configuration
        input_folder = config.get("input_folder_path")
        output_dir = config.get("output_dir")
        req_filename = config.get("req_filename", "reqs_to_use.xlsx")
        req_sheet_name = config.get("req_sheet_name", "005")

        # Resolve paths
        abs_input_folder = resolve_path(input_folder)
        abs_output_dir = resolve_path(output_dir)
        abs_excel_path = os.path.join(abs_input_folder, req_filename)

        # Log path information
        print(f"[LOG] Input folder (relative): {input_folder}")
        print(f"[LOG] Input folder (absolute): {abs_input_folder}")
        print(f"[LOG] Excel file path: {abs_excel_path}")
        print(f"[LOG] Sheet name: {req_sheet_name}")
        print(f"[LOG] Output dir (absolute): {abs_output_dir}\n")

        try:
            # Validate input file exists
            if not os.path.exists(abs_excel_path):
                error_msg = f"File not found: {abs_excel_path}"
                print(f"[ERROR] {error_msg}\n")
                errors.append(error_msg)
                state["errors"] = errors
                return state

            # Read Excel file
            print(f"[LOG] Reading Excel file...\n")
            df = pd.read_excel(abs_excel_path, sheet_name=req_sheet_name)

            print(f"[LOG] Total rows in sheet: {len(df)}")
            print(f"[LOG] Columns: {list(df.columns)}\n")

            # Extract functional requirements
            print(f"[LOG] Scanning rows using configured keywords...")
            print(f"[LOG] Keywords: {FUNCTIONAL_REQ_KEYWORDS}")
            print(f"[LOG] Matching: case_sensitive={KEYWORD_MATCHING_CONFIG['case_sensitive']}, "
                  f"exact_match={KEYWORD_MATCHING_CONFIG['exact_match']}, "
                  f"word_boundaries={KEYWORD_MATCHING_CONFIG['word_boundaries']}\n")

            for idx, row in df.iterrows():
                is_functional = False
                matched_column = None

                # Check each cell in row
                for col_name, cell_value in row.items():
                    if pd.isna(cell_value):
                        continue

                    # Use config-based matching
                    if is_functional_requirement(str(cell_value)):
                        is_functional = True
                        matched_column = col_name
                        print(f"[LOG] Row {idx}: MATCH in column '{col_name}' = '{cell_value}'")
                        break

                # Extract matching row
                if is_functional:
                    row_dict = row.to_dict()
                    row_dict = {k: (None if pd.isna(v) else v) for k, v in row_dict.items()}

                    requirement = {
                        "row_index": int(idx),
                        "data": row_dict,
                        "type": "Functional"
                    }
                    requirements.append(requirement)
                    print(f"      → Extracted: {row_dict}\n")

            # Summary
            print(f"\n[SUMMARY] Total functional requirements extracted: {len(requirements)}\n")

            # Extract verification criteria from each requirement
            print(f"\n[LOG] Extracting verification criteria from requirements...\n")
            for req in requirements:
                criteria = extract_verification_criteria(req["data"])
                req["verification_criteria"] = criteria
                if criteria:
                    print(f"[LOG] Req {req.get('row_index')}: Found criteria: {criteria[:100]}...\n")

            # Create output directory
            ensure_directory_exists(abs_output_dir)
            print(f"[LOG] Output directory verified: {abs_output_dir}\n")

            # Save to JSON
            timestamp = state["timestamp"]
            json_file = os.path.join(abs_output_dir, f"requirements_{timestamp}.json")

            output_data = {
                "metadata": {
                    "total_requirements": len(requirements),
                    "timestamp": timestamp,
                    "input_file": abs_excel_path,
                    "sheet_name": req_sheet_name,
                    "output_file": json_file
                },
                "requirements": requirements
            }

            with open(json_file, 'w') as f:
                json.dump(output_data, f, indent=2)

            print(f"[SUCCESS] Requirements saved to: {json_file}")
            print(f"[LOG] File size: {os.path.getsize(json_file)} bytes\n")

            # Generate test patterns for requirements with verification criteria
            print(f"\n[LOG] Generating test patterns using LLM...\n")
            reference_format = """{
  "test_cases": [
    {
      "test_case_no": 1,
      "preconditions": {
        "truck_size": "1t",
        "discharge_capacity": "25%",
        "power_control_mode": "P",
        "direction_switch": "FWD",
        "load_capacity": "NL"
      },
      "actions": {
        "option_set": "Enabled",
        "slope_angle": "0 deg"
      },
      "expected_result": "Description"
    }
  ]
}"""

            test_patterns_data = {}
            for req in requirements:
                if req.get("verification_criteria"):
                    try:
                        req_id = req["data"].get("REQ_ID", f"REQ_{req['row_index']}")
                        req_desc = req["data"].get("Description", "")
                        criteria = req["verification_criteria"]

                        print(f"[LOG] Generating test pattern for {req_id}...")

                        # Prepare prompt for LLM
                        prompt = prepare_test_pattern_prompt(req_desc, criteria, reference_format)

                        # Call LLM for test pattern generation using configured model
                        llm = get_llm()
                        print(f"[LOG] Calling LLM: {llm.model_name}...")

                        # Invoke LLM with the prompt
                        response = llm.invoke(prompt)

                        # Parse response - LangChain returns content directly
                        test_pattern_json = response.content
                        test_patterns = parse_test_patterns_json(test_pattern_json)
                        test_patterns_data[req_id] = test_patterns

                        print(f"[LOG] Generated {len(test_patterns.get('test_cases', []))} test cases for {req_id}\n")

                    except Exception as lm_error:
                        print(f"[WARNING] Failed to generate test patterns for {req_id}: {str(lm_error)}\n")

            # Save test patterns to Excel if generated
            if test_patterns_data:
                try:
                    from openpyxl import Workbook

                    excel_output = os.path.join(abs_output_dir, f"test_patterns_{timestamp}.xlsx")

                    # Create workbook
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Test Pattern"

                    # Add header
                    headers = [
                        "Requirement ID", "Test Case No", "Truck Size", "Discharge %",
                        "Power Mode", "Direction", "Load Cap", "Option Set",
                        "Slope Angle", "Expected Result"
                    ]

                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    # Add test cases
                    row = 2
                    for req_id, test_patterns in test_patterns_data.items():
                        for tc in test_patterns.get("test_cases", []):
                            ws.cell(row=row, column=1, value=req_id)
                            ws.cell(row=row, column=2, value=tc.get("test_case_no", ""))
                            ws.cell(row=row, column=3, value=tc.get("preconditions", {}).get("truck_size", ""))
                            ws.cell(row=row, column=4, value=tc.get("preconditions", {}).get("discharge_capacity", ""))
                            ws.cell(row=row, column=5, value=tc.get("preconditions", {}).get("power_control_mode", ""))
                            ws.cell(row=row, column=6, value=tc.get("preconditions", {}).get("direction_switch", ""))
                            ws.cell(row=row, column=7, value=tc.get("preconditions", {}).get("load_capacity", ""))
                            ws.cell(row=row, column=8, value=tc.get("actions", {}).get("option_set", ""))
                            ws.cell(row=row, column=9, value=tc.get("actions", {}).get("slope_angle", ""))
                            ws.cell(row=row, column=10, value=tc.get("expected_result", ""))
                            row += 1

                    # Auto-adjust column widths
                    for col in ws.columns:
                        max_length = 0
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

                    # Save workbook
                    wb.save(excel_output)
                    print(f"[SUCCESS] Test patterns saved to: {excel_output}\n")

                except Exception as excel_error:
                    print(f"[WARNING] Failed to save Excel: {str(excel_error)}\n")

        except Exception as e:
            error_msg = f"Error during extraction: {str(e)}"
            print(f"[ERROR] {error_msg}\n")
            errors.append(error_msg)

        # Update state
        state["requirements"] = requirements
        state["errors"] = errors

        # Print completion status
        print(f"{'='*80}")
        print(f"NODE 1 COMPLETED")
        print(f"  Status: {'SUCCESS' if not errors else 'FAILED'}")
        print(f"  Requirements extracted: {len(requirements)}")
        print(f"  Errors: {len(errors)}")
        print(f"{'='*80}\n")

        return state
