"""Create Command List file for Node 2 testing"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import os

output_dir = "files/test@tataelxsi.co.in/swe6_5/V1.0/SWE6/input"
os.makedirs(output_dir, exist_ok=True)

output_file = os.path.join(output_dir, "Command_List.xlsx")

# Create workbook
wb = openpyxl.Workbook()

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

print("[LOG] Creating Command List sheet...")
# Create Command List sheet
cmd_ws = wb.create_sheet("Command List")

# Headers: Column A is Signal Name, columns B-I are command details
headers = [
    "Signal Name",                           # A
    "Command Type",                          # B
    "Command Code",                          # C
    "Data Length",                           # D
    "Validation Method",                     # E
    "Response Expected",                     # F
    "Timeout (ms)",                          # G
    "Error Handling",                        # H
    "Notes"                                  # I
]

for col_idx, header in enumerate(headers, 1):
    cell = cmd_ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

# Add command data for the signals we defined in Master Comm Matrix
command_data = [
    [
        "Drv1_Rx1_CmdSpd",
        "Speed Control",
        "0x0001",
        "2 bytes",
        "CRC Check",
        "Yes",
        "100",
        "Retry 3 times",
        "Motor RPM command"
    ],
    [
        "Drv1_Rx1_CmdTorque",
        "Torque Control",
        "0x0002",
        "2 bytes",
        "CRC Check",
        "Yes",
        "100",
        "Retry 3 times",
        "Motor torque command"
    ],
    [
        "Drv1_Rx1_CmdDec",
        "Deceleration",
        "0x0003",
        "1 byte",
        "CRC Check",
        "Yes",
        "150",
        "Immediate stop",
        "Controlled deceleration"
    ],
    [
        "Status_PowerOn",
        "Status Query",
        "0x0010",
        "1 byte",
        "Parity",
        "Yes",
        "50",
        "Log and continue",
        "Power status check"
    ],
    [
        "Status_Error",
        "Error Query",
        "0x0011",
        "1 byte",
        "Parity",
        "Yes",
        "50",
        "Log and alert",
        "Error status check"
    ],
]

for row_idx, row_data in enumerate(command_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cmd_ws.cell(row=row_idx, column=col_idx, value=value)

# Adjust column widths
for column in cmd_ws.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    cmd_ws.column_dimensions[column_letter].width = min(adjusted_width, 50)

# Save workbook
wb.save(output_file)
print(f"[SUCCESS] Command List file created: {output_file}")
print(f"\nCommand List sheet structure:")
print(f"  Rows: {len(list(cmd_ws.rows))}")
print(f"  Columns (A-I):")
for col_idx in range(1, 10):
    cell_value = cmd_ws.cell(row=1, column=col_idx).value
    print(f"    Column {get_column_letter(col_idx)}: {cell_value}")
